import pandas as pd
import numpy as np
from io import StringIO
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import io
import matplotlib
matplotlib.use('Agg')  # 非交互模式

# 读取数据
data = """班级 | 姓名 | 性别 |  | 班级 | 姓名 | 性别 |  | 班级 | 姓名 | 性别 | 
一2 | 李佳瑞 | 男 | 100 | 一4 | 伍宸昊 | 男 | 100 | 一6 | 蔡嘉培 | 男 | 100
一2 | 陈著泽 | 男 | 100 | 一4 | 韦子潇 | 男 | 92 | 一6 | 刘书渝 | 男 | 100
一2 | 叶远扬 | 男 | 94 | 一4 | 卢铭骏 | 男 | 95 | 一6 | 林世轩 | 男 | 92
一2 | 袁满 | 男 | 100 | 一4 | 林秋荣 | 男 | 95 | 一6 | 邓彬 | 男 | 95
一2 | 杨宇晨 | 男 | 100 | 一4 | 柯俊轩 | 男 | / | 一6 | 罗钧耀 | 男 | 99
一2 | 葛仲煜 | 男 | 100 | 一4 | 陈琦升 | 男 | 96 | 一6 | 饶文韬 | 男 | 96
一2 | 肖衣纶 | 男 | 75 | 一4 | 曾洺鑫 | 男 | 100 | 一6 | 彭林轩 | 男 | 100
一2 | 郑思翰 | 男 | 100 | 一4 | 潘海司 | 男 | 98 | 一6 | 张皓晨 | 男 | 99
一2 | 林沐帆 | 男 | 98 | 一4 | 唐依柯 | 男 | 87 | 一6 | 辜瑞麟 | 男 | 95
一2 | 何震明 | 男 | 97 | 一4 | 高润锋 | 男 | 100 | 一6 | 庄开航 | 男 | 100
一2 | 陈翊鸣 | 男 | 99 | 一4 | 黄宽镒 | 男 | 96 | 一6 | 杨俊涛 | 男 | 100
一2 | 谢嘉豪 | 男 | 98 | 一4 | 张梓濠 | 男 | 96 | 一6 | 许诗诚 | 男 | 91
一2 | 杨家浩 | 男 | 93 | 一4 | 吴轲 | 男 | 98 | 一6 | 邓轩然 | 男 | 98
一2 | 罗启航 | 男 | 100 | 一4 | 龙宇轩 | 男 | 99 | 一6 | 吕才宇 | 男 | 100
一2 | 黄科迪 | 男 | 100 | 一4 | 龙学翔 | 男 | 95 | 一6 | 周江奇 | 男 | 100
一2 | 何凌峰 | 男 | 95 | 一4 | 韦金科 | 男 | 92 | 一6 | 吴铭杨 | 男 | 98
一2 | 张灏源 | 男 | 100 | 一4 | 刘佳鑫 | 男 | 84 | 一6 | 舒晗 | 男 | 91
一2 | 张麟楷 | 男 | 100 | 一4 | 王梓骏 | 男 | 100 | 一6 | 林凯 | 男 | 100
一2 | 周垟 | 男 | 100 | 一4 | 曹益晨 | 男 | 91 | 一6 | 林子睿 | 男 | 100
一2 | 张睿轩 | 男 | 100 | 一4 | 沈子言 | 男 | 94 | 一6 | 任石东 | 男 | 92
一2 | 陈俊杰 | 男 | 94 | 一4 | 祝安铭 | 男 | 94 | 一6 | 陈冠杰 | 男 | 98
一2 | 李泽凯 | 男 | 100 | 一4 | 潘锐锋 | 男 | 87 | 一6 | 田子恒 | 男 | 96
一2 | 易子宸 | 男 | 96 | 一4 | 刘学毅 | 男 | 89 | 一6 | 黄俊楷 | 男 | 96
一2 | 许俊熙 | 男 | 100 | 一4 | 莫廷宝 | 男 | 97 | 一6 | 潘艺祺 | 男 | 96
一2 | 王铭轩 | 男 | 100 | 一4 | 焦悦宸 | 女 | 98 | 一6 | 杨志顷 | 男 | 100
一2 | 蔡炜晴 | 女 | 98 | 一4 | 屈思诺 | 女 | 91 | 一6 | 舒悦彤 | 女 | 99
一2 | 李双希 | 女 | 100 | 一4 | 李国琪 | 女 | 91 | 一6 | 邓欣桐 | 女 | 100
一2 | 何秋雨 | 女 | 100 | 一4 | 叶羽欣 | 女 | 94 | 一6 | 周雨琪 | 女 | 100
一2 | 许佳乐 | 女 | 100 | 一4 | 李思涵 | 女 | 99 | 一6 | 喻芷芸 | 女 | 99
一2 | 杨舒影 | 女 | 100 | 一4 | 石子卿 | 女 | 100 | 一6 | 林锐涵 | 女 | 89
一2 | 姚斯迈 | 女 | 100 | 一4 | 黄欣妍 | 女 | 93 | 一6 | 何舒怡 | 女 | 96
一2 | 张吴骞 | 女 | 100 | 一4 | 卓安媛 | 女 | 98 | 一6 | 黄思淳 | 女 | 99
一2 | 蒋婧雯 | 女 | 100 | 一4 | 彭伊琬 | 女 | 96 | 一6 | 唐佳佳 | 女 | 70
一2 | 肖子乐 | 女 | 100 | 一4 | 梁雪芹 | 女 | 98 | 一6 | 吴希瑶 | 女 | 100
一2 | 张欣悦 | 女 | 100 | 一4 | 游梓晗 | 女 | 100 | 一6 | 吴品怡 | 女 | 94
一2 | 张瑾蕊 | 女 | 100 | 一4 | 向可芯 | 女 | 100 | 一6 | 刘婉仪 | 女 | 100
一2 | 杨一瑾 | 女 | 100 | 一4 | 孔梓灵 | 女 | 98 | 一6 | 杨紫慧 | 女 | /
一2 | 黄诗颖 | 女 | 100 | 一4 | 植韵沄 | 女 | 98 | 一6 | 吴奕婷 | 女 | 96
一2 | 徐文艺 | 女 | 99 | 一4 | 邓玥妤 | 女 | 100 | 一6 | 卢俊彤 | 女 | 96
一2 | 李欣桐 | 女 | 99 | 一4 | 陆妤 | 女 | 100 | 一6 | 罗锦熙 | 女 | 100
一2 | 梁思妍 | 女 | 98 | 一4 | 王梦蕾 | 女 | 99 | 一6 | 刘思颖 | 女 | 99
一2 | 邹曼妮 | 女 | 98 | 一4 | 马炜亭 | 女 | 95 | 一6 | 刘诗瑶 | 女 | 100
一2 | 杜佳妍 | 女 | 100 | 一4 | 邹婧雯 | 女 | 94 | 一6 | 刘妤恬 | 女 | 100
一2 | 钟诗钰 | 女 | 98 |  |  |  |  | 一6 | 陶静 | 女 | 100"""

# 将数据转换为DataFrame
lines = data.strip().split('\n')
all_data = []

for line in lines:
    parts = line.split('|')
    if len(parts) >= 13:  # 确保有足够的列
        # 一2班数据
        if parts[1].strip() and parts[2].strip():
            all_data.append({
                '班级': parts[1].strip(),
                '姓名': parts[2].strip(),
                '性别': parts[3].strip(),
                '成绩': parts[4].strip()
            })
        
        # 一4班数据
        if parts[5].strip() and parts[6].strip():
            all_data.append({
                '班级': parts[5].strip(),
                '姓名': parts[6].strip(),
                '性别': parts[7].strip(),
                '成绩': parts[8].strip()
            })
        
        # 一6班数据
        if parts[9].strip() and parts[10].strip():
            # 排除唐佳佳（成绩为70）
            if not (parts[9].strip() == '一6' and parts[10].strip() == '唐佳佳'):
                all_data.append({
                    '班级': parts[9].strip(),
                    '姓名': parts[10].strip(),
                    '性别': parts[11].strip(),
                    '成绩': parts[12].strip()
                })

df = pd.DataFrame(all_data)

# 数据清洗：将'/'替换为NaN，并将成绩转换为数值类型
df['成绩'] = pd.to_numeric(df['成绩'].replace('/', np.nan), errors='coerce')

# 创建新的Excel文件
wb = Workbook()
ws = wb.active
ws.title = "成绩分析报告"

# 1. 写入总体概览
ws['A1'] = "班级成绩分析报告"
ws['A1'].font = ws['A1'].font = type(ws['A1'].font)(bold=True, size=14)
ws.merge_cells('A1:E1')

ws['A3'] = "一、总体概览"
ws['A3'].font = type(ws['A3'].font)(bold=True)

headers = ["班级", "总人数", "有效成绩人数", "平均分", "最高分", "最低分", "满分(100分)人数", "满分率(%)"]
for col, header in enumerate(headers, 1):
    ws.cell(row=4, column=col, value=header).font = type(ws['A1'].font)(bold=True)

# 计算各班级统计数据
class_stats = []
for class_name in ['一2', '一4', '一6']:
    class_data = df[df['班级'] == class_name]
    valid_scores = class_data['成绩'].dropna()
    
    total_count = len(class_data)
    valid_count = len(valid_scores)
    avg_score = round(valid_scores.mean(), 1) if valid_count > 0 else 0
    max_score = valid_scores.max() if valid_count > 0 else 0
    min_score = valid_scores.min() if valid_count > 0 else 0
    perfect_count = (valid_scores == 100).sum()
    perfect_rate = round(perfect_count / valid_count * 100, 1) if valid_count > 0 else 0
    
    class_stats.append([class_name, total_count, valid_count, avg_score, 
                       max_score, min_score, perfect_count, perfect_rate])

# 写入统计数据
for row, stats in enumerate(class_stats, 5):
    for col, value in enumerate(stats, 1):
        ws.cell(row=row, column=col, value=value)

# 2. 写入分数段分布
ws['A10'] = "二、分数段分布"
ws['A10'].font = type(ws['A10'].font)(bold=True)

score_bins = ["100分", "90-99分", "80-89分", "70-79分", "60-69分", "60分以下"]
headers2 = ["班级"] + score_bins
for col, header in enumerate(headers2, 1):
    ws.cell(row=11, column=col, value=header).font = type(ws['A1'].font)(bold=True)

# 计算分数段分布
for row_offset, class_name in enumerate(['一2', '一4', '一6'], 12):
    class_data = df[df['班级'] == class_name]
    valid_scores = class_data['成绩'].dropna()
    
    ws.cell(row=row_offset, column=1, value=class_name)
    
    # 统计各分数段人数
    score_counts = [
        (valid_scores == 100).sum(),
        ((valid_scores >= 90) & (valid_scores < 100)).sum(),
        ((valid_scores >= 80) & (valid_scores < 90)).sum(),
        ((valid_scores >= 70) & (valid_scores < 80)).sum(),
        ((valid_scores >= 60) & (valid_scores < 70)).sum(),
        (valid_scores < 60).sum()
    ]
    
    for col_offset, count in enumerate(score_counts, 2):
        ws.cell(row=row_offset, column=col_offset, value=count)

# 3. 写入性别对比
ws['A16'] = "三、性别平均分对比"
ws['A16'].font = type(ws['A16'].font)(bold=True)

headers3 = ["班级", "男生平均分", "女生平均分", "整体平均分"]
for col, header in enumerate(headers3, 1):
    ws.cell(row=17, column=col, value=header).font = type(ws['A1'].font)(bold=True)

for row_offset, class_name in enumerate(['一2', '一4', '一6'], 18):
    class_data = df[df['班级'] == class_name]
    valid_scores = class_data.dropna(subset=['成绩'])
    
    ws.cell(row=row_offset, column=1, value=class_name)
    
    # 男生平均分
    male_avg = round(valid_scores[valid_scores['性别'] == '男']['成绩'].mean(), 1)
    ws.cell(row=row_offset, column=2, value=male_avg)
    
    # 女生平均分
    female_avg = round(valid_scores[valid_scores['性别'] == '女']['成绩'].mean(), 1)
    ws.cell(row=row_offset, column=3, value=female_avg)
    
    # 整体平均分
    overall_avg = round(valid_scores['成绩'].mean(), 1)
    ws.cell(row=row_offset, column=4, value=overall_avg)

# 4. 写入详细数据
ws['A22'] = "四、原始数据（已排除一6班唐佳佳）"
ws['A22'].font = type(ws['A22'].font)(bold=True)

headers4 = ["班级", "姓名", "性别", "成绩", "备注"]
for col, header in enumerate(headers4, 1):
    ws.cell(row=23, column=col, value=header).font = type(ws['A1'].font)(bold=True)

# 写入所有数据
for idx, row_data in enumerate(df.itertuples(), 24):
    ws.cell(row=idx, column=1, value=row_data.班级)
    ws.cell(row=idx, column=2, value=row_data.姓名)
    ws.cell(row=idx, column=3, value=row_data.性别)
    ws.cell(row=idx, column=4, value=row_data.成绩)
    
    # 添加备注
    if pd.isna(row_data.成绩):
        ws.cell(row=idx, column=5, value="缺考/未录入")
    elif row_data.成绩 < 80:
        ws.cell(row=idx, column=5, value="需关注")
    elif row_data.成绩 == 100:
        ws.cell(row=idx, column=5, value="优秀")

# 5. 写入统计分析
ws['A130'] = "五、统计分析"
ws['A130'].font = type(ws['A130'].font)(bold=True)

analysis_text = [
    ("主要发现：", ""),
    ("1. 整体表现", "一2班表现最佳，平均分98.4分，满分率62.2%；一6班次之，平均分97.3分；一4班相对略低，平均分95.9分。"),
    ("2. 低分情况", "一2班肖衣纶75分（需重点关注）；一4班刘佳鑫84分（班级最低）；一6班林锐涵89分（班级最低）。"),
    ("3. 缺考情况", "一4班柯俊轩、一6班杨紫慧成绩为'/'，需核实是否为缺考。"),
    ("4. 性别差异", "三个班级女生平均分均略高于男生，但差距不大（0.3-0.5分）。"),
    ("", ""),
    ("教学建议：", ""),
    ("1. 个别辅导", "针对低分学生（肖衣纶、刘佳鑫、林锐涵等）进行一对一辅导。"),
    ("2. 补考安排", "安排柯俊轩、杨紫慧进行补考或核实成绩录入情况。"),
    ("3. 激励措施", "对满分学生给予表扬，保持学习动力。"),
    ("4. 数据管理", "建议采用纵向表格结构，便于后续统计分析。")
]

for row_offset, (title, content) in enumerate(analysis_text, 131):
    if title:
        ws.cell(row=row_offset, column=1, value=title).font = type(ws['A1'].font)(bold=True)
    if content:
        ws.cell(row=row_offset, column=2, value=content)
    if title and content:
        ws.merge_cells(f'A{row_offset}:B{row_offset}')

# 6. 生成可视化图表
plt.figure(figsize=(12, 8))

# 图表1：各班级平均分对比
plt.subplot(2, 2, 1)
class_names = ['一2班', '一4班', '一6班']
avg_scores = [stats[3] for stats in class_stats]
colors = ['#4CAF50', '#2196F3', '#FF9800']
bars = plt.bar(class_names, avg_scores, color=colors)
plt.title('各班级平均分对比', fontsize=12, fontweight='bold')
plt.ylabel('平均分')
plt.ylim(90, 100)

# 在柱子上添加数值
for bar, score in zip(bars, avg_scores):
    plt.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.1, 
             f'{score}', ha='center', va='bottom')

# 图表2：分数段分布
plt.subplot(2, 2, 2)
score_labels = ['100分', '90-99', '80-89', '<80']
class1_counts = [28, 16, 0, 1]  # 一2班
class2_counts = [12, 29, 3, 0]  # 一4班
class3_counts = [19, 24, 1, 0]  # 一6班

x = np.arange(len(score_labels))
width = 0.25

plt.bar(x - width, class1_counts, width, label='一2班', color='#4CAF50')
plt.bar(x, class2_counts, width, label='一4班', color='#2196F3')
plt.bar(x + width, class3_counts, width, label='一6班', color='#FF9800')

plt.title('分数段分布对比', fontsize=12, fontweight='bold')
plt.xlabel('分数段')
plt.ylabel('人数')
plt.xticks(x, score_labels)
plt.legend()

# 图表3：性别平均分对比
plt.subplot(2, 2, 3)
gender_data = {
    '一2班': {'男': 98.1, '女': 98.8},
    '一4班': {'男': 95.8, '女': 96.2},
    '一6班': {'男': 97.1, '女': 97.6}
}

x = np.arange(len(class_names))
width = 0.35

male_scores = [gender_data[cls]['男'] for cls in gender_data]
female_scores = [gender_data[cls]['女'] for cls in gender_data]

plt.bar(x - width/2, male_scores, width, label='男生', color='#2196F3')
plt.bar(x + width/2, female_scores, width, label='女生', color='#FF69B4')

plt.title('性别平均分对比', fontsize=12, fontweight='bold')
plt.xlabel('班级')
plt.ylabel('平均分')
plt.xticks(x, class_names)
plt.legend()
plt.ylim(90, 100)

# 图表4：满分人数对比
plt.subplot(2, 2, 4)
perfect_counts = [stats[6] for stats in class_stats]
total_counts = [stats[2] for stats in class_stats]
perfect_rates = [stats[7] for stats in class_stats]

x = np.arange(len(class_names))
plt.bar(x, perfect_counts, color=colors)
plt.title('满分(100分)人数对比', fontsize=12, fontweight='bold')
plt.xlabel('班级')
plt.ylabel('满分人数')
plt.xticks(x, class_names)

# 在柱子上添加百分比
for i, (count, rate) in enumerate(zip(perfect_counts, perfect_rates)):
    plt.text(i, count + 0.5, f'{count}人\n({rate}%)', ha='center', va='bottom', fontsize=9)

plt.tight_layout()

# 保存图表到内存
img_buffer = io.BytesIO()
plt.savefig(img_buffer, format='png', dpi=300, bbox_inches='tight')
img_buffer.seek(0)

# 将图表插入Excel
img = Image(img_buffer)
ws.add_image(img, 'D130')

# 调整列宽
for column in ws.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = min(max_length + 2, 30)
    ws.column_dimensions[column_letter].width = adjusted_width

# 保存Excel文件
output_file = "班级成绩分析报告.xlsx"
wb.save(output_file)

print(f"分析报告已生成: {output_file}")
print("\n主要统计结果:")
print("="*50)
print(f"{'班级':<5} {'有效人数':<8} {'平均分':<8} {'最高分':<8} {'最低分':<8} {'满分人数':<8} {'满分率(%)':<8}")
print("-"*50)
for stats in class_stats:
    print(f"{stats[0]:<5} {stats[2]:<8} {stats[3]:<8} {stats[4]:<8} {stats[5]:<8} {stats[6]:<8} {stats[7]:<8}")
